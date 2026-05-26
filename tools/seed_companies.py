"""
Reads Company List.xlsx (read-only) and writes companies.json.
Run once to bootstrap; after that manage companies via the Streamlit UI or by editing companies.json.
"""
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent

try:
    import openpyxl
except ImportError:
    sys.exit("Run: pip install openpyxl")

XLSX_PATH = ROOT / "Company List.xlsx"
OUTPUT_PATH = ROOT / "companies.json"

# Normalize raw platform strings → canonical keys
PLATFORM_MAP = {
    "greenhouse (api)": "greenhouse",
    "greenhouse":       "greenhouse",
    "lever (api)":      "lever",
    "lever":            "lever",
    "ashby (api)":      "ashby",
    "ashby":            "ashby",
    "jobvite (api)":    "jobvite",
    "jobvite":          "jobvite",
    "smartrecruiters (api)": "smartrecruiters",
    "smartrecruiters":  "smartrecruiters",
    "workable (api)":   "workable",
    "workable":         "workable",
    "pinpoint (api)":   "pinpoint",
    "rippling (api)":   "rippling",
    "workday":          "workday",
    "icims":            "icims",
    "adp":              "adp",
    "csod":             "csod",
    "dayforce":         "dayforce",
    "successfactors":   "successfactors",
    "taleo":            "taleo",
    "ukg":              "ukg",
    "breezy":           "breezy",
    "phenom":           "phenom",
    "silkroad":         "silkroad",
    "paycom":           "paycom",
    "paylocity":        "paylocity",
    "oracle":           "oracle",
    "ibm":              "ibm",
    "linkedin":         "linkedin",
    "indeed":           "indeed",
    "google":           "google",
    "notion":           "notion",
    "custom":           "custom",
    "consider":         "custom",
    "rippelhire":       "custom",
    "gupy":             "custom",
    "n/a":              None,
    "none":             None,
}

# Platforms with free unauthenticated REST APIs
API_PLATFORMS = {"greenhouse", "lever", "ashby", "workable", "smartrecruiters"}

# URL patterns for extracting ATS slugs
SLUG_PATTERNS = {
    "greenhouse": [
        r"boards\.greenhouse\.io/([^/?#\s]+)",
        r"job-boards\.greenhouse\.io/([^/?#\s]+)",
        r"boards\.eu\.greenhouse\.io/([^/?#\s]+)",
    ],
    "lever": [
        r"jobs\.lever\.co/([^/?#\s]+)",
    ],
    "ashby": [
        r"jobs\.ashbyhq\.com/([^/?#\s]+)",
    ],
    "workable": [
        r"apply\.workable\.com/([^/?#\s]+)",
        r"jobs\.workable\.com/([^/?#\s]+)",
    ],
    "smartrecruiters": [
        r"jobs\.smartrecruiters\.com/([^/?#\s]+)",
        r"careers\.smartrecruiters\.com/([^/?#\s]+)",
    ],
    "jobvite": [
        r"jobs\.jobvite\.com/([^/?#\s]+)",
    ],
    "pinpoint": [
        r"([^.]+)\.pinpointhq\.com",
    ],
}

CHECK_INTERVALS = {
    "greenhouse": 1, "lever": 1, "ashby": 1,
    "workable": 1, "smartrecruiters": 1, "jobvite": 2,
}
DEFAULT_INTERVAL = 6


def extract_slug(platform: str, url: str) -> str | None:
    patterns = SLUG_PATTERNS.get(platform, [])
    for pat in patterns:
        m = re.search(pat, url or "", re.IGNORECASE)
        if m:
            return m.group(1).rstrip("/")
    return None


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip())


def seed(xlsx_path: Path = XLSX_PATH, output_path: Path = OUTPUT_PATH, overwrite: bool = False):
    if output_path.exists() and not overwrite:
        print(f"  {output_path.name} already exists. Use --overwrite to replace it.")
        return

    print(f"Reading {xlsx_path.name} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Build header → column-index map
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    name_i  = col["Organization Name"]
    plat_i  = col["Platform"]
    care_i  = col["Careers Page"]
    mon_i   = col["Monitoring Link"]
    notes_i = col.get("Careers Page Notes")

    companies = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = normalize_name(str(row[name_i] or ""))
        if not name:
            continue

        raw_platform = str(row[plat_i] or "").strip()
        platform = PLATFORM_MAP.get(raw_platform.lower())

        # Prefer Monitoring Link (more specific URL); fall back to Careers Page
        url = str(row[mon_i] or "").strip() or str(row[care_i] or "").strip()
        notes = str(row[notes_i] or "").strip() if notes_i is not None else ""

        if not url:
            skipped += 1
            continue

        slug = extract_slug(platform, url) if platform else None
        use_api = platform in API_PLATFORMS and slug is not None
        interval = CHECK_INTERVALS.get(platform, DEFAULT_INTERVAL)

        companies.append({
            "name":                  name,
            "platform":              platform,
            "ats_slug":              slug,
            "careers_url":           url,
            "scraping_method":       "api" if use_api else "playwright",
            "check_interval_hours":  interval,
            "active":                True,
            "notes":                 notes,
        })

    wb.close()

    output_path.write_text(json.dumps(companies, indent=2, ensure_ascii=False), encoding="utf-8")

    api_count = sum(1 for c in companies if c["scraping_method"] == "api")
    pw_count  = len(companies) - api_count
    print(f"  Wrote {len(companies)} companies to {output_path.name}")
    print(f"  API-ready: {api_count}  |  Playwright: {pw_count}  |  Skipped (no URL): {skipped}")


if __name__ == "__main__":
    overwrite = "--overwrite" in sys.argv
    seed(overwrite=overwrite)
